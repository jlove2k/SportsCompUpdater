from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def load_workbook_for_update(path: Path, sheet_name: str | None = None) -> tuple[Workbook, Worksheet]:
    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")

    workbook = load_workbook(path)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            available = ", ".join(workbook.sheetnames)
            raise ValueError(f"Worksheet '{sheet_name}' not found. Available sheets: {available}")
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active

    return workbook, sheet


def save_workbook(workbook: Any, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)

